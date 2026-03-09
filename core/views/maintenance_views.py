"""
maintenance_views.py — Техническое обслуживание оборудования
v3.31.0

Расположение: core/views/maintenance_views.py

Подключить в core/views/__init__.py:
    from . import maintenance_views

Маршруты в core/urls.py:
    path('workspace/maintenance/', maintenance_views.maintenance_view, name='maintenance'),
    path('workspace/maintenance/<int:plan_id>/', maintenance_views.maintenance_detail_view, name='maintenance_detail'),
    path('workspace/maintenance/save-columns/', maintenance_views.save_maintenance_columns, name='save_maintenance_columns'),
    path('workspace/maintenance/save-column-widths/', maintenance_views.save_maintenance_column_widths, name='save_maintenance_column_widths'),
    path('workspace/maintenance/export/', maintenance_views.export_maintenance_xlsx, name='export_maintenance_xlsx'),
"""

import json
from datetime import date
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import connection
from django.http import HttpResponseForbidden, JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from urllib.parse import urlencode

from core.permissions import PermissionChecker
from core.models import Laboratory

MAINTENANCE_ITEMS_PER_PAGE = 50
MAINTENANCE_PER_PAGE_OPTIONS = [50, 100, 200]
LOG_ITEMS_PER_PAGE = 50

LOG_STATUS_CHOICES = [
    ('PLANNED',   'Запланировано'),
    ('COMPLETED', 'Выполнено'),
    ('OVERDUE',   'Просрочено'),
    ('CANCELLED', 'Отменено'),
]
LOG_STATUS_LABELS = dict(LOG_STATUS_CHOICES)

# ═════════════════════════════════════════════════════════════════
# Столбцы для планов ТО ⭐ v3.31.0
# ═════════════════════════════════════════════════════════════════

MAINTENANCE_DISPLAYABLE_COLUMNS = [
    ('accounting_number',    'Инв. номер'),
    ('equipment_type',       'Тип'),
    ('laboratory_name',      'Лаборатория'),
    ('name',                 'Вид обслуживания'),
    ('frequency_display',    'Периодичность'),
    ('frequency_condition',  'Условие'),
    ('next_due_date',        'Следующая дата'),
    ('notes',                'Примечания'),
]

MAINTENANCE_COLUMNS_DICT = {code: name for code, name in MAINTENANCE_DISPLAYABLE_COLUMNS}

DEFAULT_MAINTENANCE_COLUMNS = [
    'accounting_number', 'equipment_type', 'laboratory_name',
    'name', 'frequency_display', 'frequency_condition',
    'next_due_date', 'notes',
]


def _get_maintenance_user_columns(user):
    """Возвращает список выбранных столбцов для планов ТО."""
    prefs = user.ui_preferences or {}
    saved = prefs.get('journal_columns', {}).get('MAINTENANCE')
    if saved:
        all_codes = {code for code, _ in MAINTENANCE_DISPLAYABLE_COLUMNS}
        return [c for c in saved if c in all_codes]
    return list(DEFAULT_MAINTENANCE_COLUMNS)


# ─────────────────────────────────────────────────────────────
# Утилиты
# ─────────────────────────────────────────────────────────────

def _fetchall(sql, params=None):
    with connection.cursor() as cur:
        cur.execute(sql, params or [])
        cols = [col[0] for col in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


def _fetchone(sql, params=None):
    with connection.cursor() as cur:
        cur.execute(sql, params or [])
        row = cur.fetchone()
        if row is None:
            return None
        cols = [col[0] for col in cur.description]
        return dict(zip(cols, row))


def _build_frequency_display(plan):
    """Склеивает три поля периодичности в одну строку."""
    count        = plan.get('frequency_count')
    period_value = plan.get('frequency_period_value')
    unit         = plan.get('frequency_unit') or ''
    condition    = plan.get('frequency_condition') or ''
    is_cond      = plan.get('is_condition_based', False)

    if is_cond and not count:
        return condition or 'По условию'

    parts = []
    if count and unit and period_value:
        unit_forms = {
            'DAY':   ('день', 'дня', 'дней'),
            'WEEK':  ('неделю', 'недели', 'недель'),
            'MONTH': ('месяц', 'месяца', 'месяцев'),
            'YEAR':  ('год', 'года', 'лет'),
        }
        forms = unit_forms.get(unit, (unit,) * 3)

        def _pl(n, f1, f2, f5):
            n_abs = abs(n) % 100
            if 11 <= n_abs <= 19:
                return f5
            last = n_abs % 10
            if last == 1: return f1
            if 2 <= last <= 4: return f2
            return f5

        unit_word = _pl(period_value, *forms)

        raz_word = _pl(count, 'раз', 'раза', 'раз')

        if period_value == 1:
            if count == 1:
                parts.append(f'раз в {unit_word}')
            else:
                parts.append(f'{count} {raz_word} в {unit_word}')
        else:
            if count == 1:
                parts.append(f'раз в {period_value} {unit_word}')
            else:
                parts.append(f'{count} {raz_word} в {period_value} {unit_word}')

    if is_cond and condition:
        parts.append(f'({condition})')

    return ', '.join(parts) if parts else '—'


def _recalculate_next_due_date(plan_id):
    """
    Пересчитывает next_due_date на основе последнего выполненного обслуживания.
    Формула: performed_date + (interval / frequency_count)
    """
    with connection.cursor() as cur:
        cur.execute("""
            UPDATE equipment_maintenance_plans p
            SET next_due_date = l.performed_date + (
                CASE p.frequency_unit
                    WHEN 'DAY'   THEN make_interval(days  => p.frequency_period_value)
                    WHEN 'WEEK'  THEN make_interval(weeks => p.frequency_period_value)
                    WHEN 'MONTH' THEN make_interval(months => p.frequency_period_value)
                    WHEN 'YEAR'  THEN make_interval(years  => p.frequency_period_value)
                END / p.frequency_count
            ),
            updated_at = CURRENT_TIMESTAMP
            FROM equipment_maintenance_logs l
            WHERE l.plan_id = p.id
              AND p.id = %s
              AND l.status IN ('COMPLETED', 'OVERDUE')
              AND l.performed_date = (
                  SELECT MAX(l2.performed_date)
                  FROM equipment_maintenance_logs l2
                  WHERE l2.plan_id = p.id
                    AND l2.status IN ('COMPLETED', 'OVERDUE')
              )
        """, [plan_id])


# ─────────────────────────────────────────────────────────────
# Реестр планов ТО
# ─────────────────────────────────────────────────────────────

@login_required
def maintenance_view(request):
    if not PermissionChecker.can_view(request.user, 'MAINTENANCE', 'access'):
        messages.error(request, 'У вас нет доступа к разделу технического обслуживания')
        return redirect('workspace_home')

    can_edit = PermissionChecker.can_edit(request.user, 'MAINTENANCE', 'access')

    search       = request.GET.get('search', '').strip()
    lab_id       = request.GET.get('lab_id', '')
    overdue_only = request.GET.get('overdue_only', '')
    sort         = request.GET.get('sort', 'next_due_date')

    allowed_sorts = {
        'accounting_number', '-accounting_number',
        'equipment_type',    '-equipment_type',
        'name',              '-name',
        'next_due_date',     '-next_due_date',
    }
    if sort not in allowed_sorts:
        sort = 'next_due_date'

    sort_col = sort.lstrip('-')
    sort_dir = 'DESC' if sort.startswith('-') else 'ASC'
    sort_col_map = {
        'accounting_number': 'e.accounting_number',
        'equipment_type':    'e.equipment_type',
        'name':              'emp.name',
        'next_due_date':     'emp.next_due_date',
    }
    order_by = f"{sort_col_map.get(sort_col, 'emp.next_due_date')} {sort_dir}"

    where_clauses, params = [], []

    if search:
        where_clauses.append(
            "(e.accounting_number ILIKE %s OR emp.name ILIKE %s OR e.equipment_type ILIKE %s)"
        )
        like = f'%{search}%'
        params += [like, like, like]

    if lab_id:
        where_clauses.append("e.laboratory_id = %s")
        params.append(int(lab_id))

    if overdue_only:
        where_clauses.append("emp.next_due_date < CURRENT_DATE")

    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    rows = _fetchall(f"""
        SELECT
            emp.id,
            e.accounting_number,
            e.equipment_type,
            l.name          AS laboratory_name,
            emp.name,
            emp.frequency_count,
            emp.frequency_period_value,
            emp.frequency_unit,
            emp.frequency_condition,
            emp.next_due_date,
            emp.notes
        FROM equipment_maintenance_plans emp
        JOIN equipment e ON e.id = emp.equipment_id
        LEFT JOIN laboratories l ON l.id = e.laboratory_id
        {where_sql}
        ORDER BY {order_by}
    """, params)

    today = date.today()
    for row in rows:
        row['frequency_display'] = _build_frequency_display(row)
        nd = row.get('next_due_date')
        if nd:
            row['is_overdue'] = nd < today
            row['days_left']  = (nd - today).days
        else:
            row['is_overdue'] = False
            row['days_left']  = None

    try:
        per_page = int(request.GET.get('per_page', MAINTENANCE_ITEMS_PER_PAGE))
        if per_page not in MAINTENANCE_PER_PAGE_OPTIONS:
            per_page = MAINTENANCE_ITEMS_PER_PAGE
    except (ValueError, TypeError):
        per_page = MAINTENANCE_ITEMS_PER_PAGE

    paginator = Paginator(rows, per_page)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
    laboratories = Laboratory.objects.filter(is_active=True, department_type='LAB').order_by('name')

    # Подсчёт активных фильтров
    active_filter_count = 0
    if search: active_filter_count += 1
    if lab_id: active_filter_count += 1
    if overdue_only: active_filter_count += 1

    # Подсчёт просроченных
    overdue_count = sum(1 for r in rows if r.get('is_overdue'))

    # ─── Столбцы ⭐ v3.31.0 ───
    selected_columns = _get_maintenance_user_columns(request.user)
    visible_columns = [
        {'code': code, 'name': MAINTENANCE_COLUMNS_DICT[code]}
        for code in selected_columns
        if code in MAINTENANCE_COLUMNS_DICT
    ]
    all_available_columns = []
    for code in selected_columns:
        if code in MAINTENANCE_COLUMNS_DICT:
            all_available_columns.append({'code': code, 'name': MAINTENANCE_COLUMNS_DICT[code], 'selected': True})
    for code, _ in MAINTENANCE_DISPLAYABLE_COLUMNS:
        if code not in selected_columns:
            all_available_columns.append({'code': code, 'name': MAINTENANCE_COLUMNS_DICT[code], 'selected': False})

    # Ширины столбцов
    prefs = request.user.ui_preferences or {}
    column_widths = prefs.get('maintenance_column_widths', {})

    filter_params = {}
    if search:       filter_params['search']       = search
    if lab_id:       filter_params['lab_id']       = lab_id
    if overdue_only: filter_params['overdue_only'] = overdue_only
    if sort != 'next_due_date': filter_params['sort'] = sort
    if per_page != MAINTENANCE_ITEMS_PER_PAGE: filter_params['per_page'] = per_page

    context = {
        'page_obj':             page_obj,
        'plans':                page_obj.object_list,
        'total_count':          len(rows),
        'overdue_count':        overdue_count,
        'active_filter_count':  active_filter_count,
        'laboratories':         laboratories,
        'can_edit':             can_edit,
        'current_search':       search,
        'current_lab_id':       lab_id,
        'current_overdue_only': overdue_only,
        'current_sort':         sort,
        'filter_query':         urlencode(filter_params),
        'sort_link_params':     urlencode({k: v for k, v in filter_params.items() if k != 'sort'}),
        'per_page':             per_page,
        'per_page_options':     MAINTENANCE_PER_PAGE_OPTIONS,
        'user':                 request.user,
        'visible_columns':      visible_columns,
        'all_available_columns': all_available_columns,
        'column_widths':        json.dumps(column_widths),
    }
    return render(request, 'core/maintenance.html', context)


# ─────────────────────────────────────────────────────────────
# Детальная страница плана ТО
# ─────────────────────────────────────────────────────────────

@login_required
def maintenance_detail_view(request, plan_id):
    if not PermissionChecker.can_view(request.user, 'MAINTENANCE', 'access'):
        messages.error(request, 'У вас нет доступа к разделу технического обслуживания')
        return redirect('workspace_home')

    can_edit = PermissionChecker.can_edit(request.user, 'MAINTENANCE', 'access')

    # Шапка плана
    plan = _fetchone("""
        SELECT
            emp.id,
            e.accounting_number,
            e.name              AS equipment_name,
            e.equipment_type,
            l.name              AS laboratory_name,
            emp.name,
            emp.frequency_count,
            emp.frequency_period_value,
            emp.frequency_unit,
            emp.frequency_condition,
            emp.next_due_date,
            emp.notes
        FROM equipment_maintenance_plans emp
        JOIN equipment e ON e.id = emp.equipment_id
        LEFT JOIN laboratories l ON l.id = e.laboratory_id
        WHERE emp.id = %s
    """, [plan_id])

    if plan is None:
        messages.error(request, 'План обслуживания не найден')
        return redirect('maintenance')

    plan['frequency_display'] = _build_frequency_display(plan)
    today = date.today()
    nd = plan.get('next_due_date')
    if nd:
        plan['is_overdue'] = nd < today
        plan['days_left']  = (nd - today).days
    else:
        plan['is_overdue'] = False
        plan['days_left']  = None

    # ── POST: добавление записи ───────────────────────────────
    if request.method == 'POST':
        if not can_edit:
            return HttpResponseForbidden()

        performed_date  = request.POST.get('performed_date', '').strip()
        performed_by_id = request.POST.get('performed_by_id', '').strip() or None
        verified_date   = request.POST.get('verified_date', '').strip()   or None
        verified_by_id  = request.POST.get('verified_by_id', '').strip()  or None
        status          = request.POST.get('status', 'COMPLETED')
        log_notes       = request.POST.get('notes', '').strip()           or None

        if not performed_date:
            messages.error(request, 'Укажите дату проведения обслуживания')
        else:
            with connection.cursor() as cur:
                cur.execute("""
                    INSERT INTO equipment_maintenance_logs
                        (plan_id, performed_date, performed_by_id,
                         verified_date, verified_by_id, status, notes, created_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
                """, [
                    plan_id,
                    performed_date,
                    int(performed_by_id) if performed_by_id else None,
                    verified_date,
                    int(verified_by_id) if verified_by_id else None,
                    status,
                    log_notes,
                ])

            if status in ('COMPLETED', 'OVERDUE'):
                _recalculate_next_due_date(plan_id)

            messages.success(request, 'Запись об обслуживании добавлена')
            return redirect('maintenance_detail', plan_id=plan_id)

    # ── Журнал обслуживания ───────────────────────────────────
    logs = _fetchall("""
        SELECT
            ml.id,
            ml.performed_date,
            ml.verified_date,
            ml.status,
            ml.notes,
            pb.last_name  AS performed_last,
            pb.first_name AS performed_first,
            pb.sur_name   AS performed_sur,
            vb.last_name  AS verified_last,
            vb.first_name AS verified_first,
            vb.sur_name   AS verified_sur
        FROM equipment_maintenance_logs ml
        LEFT JOIN users pb ON pb.id = ml.performed_by_id
        LEFT JOIN users vb ON vb.id = ml.verified_by_id
        WHERE ml.plan_id = %s
        ORDER BY ml.performed_date DESC
    """, [plan_id])

    for log in logs:
        log['performed_by'] = ' '.join(filter(None, [
            log.get('performed_last'),
            log.get('performed_first'),
            log.get('performed_sur'),
        ])) or '—'
        log['verified_by'] = ' '.join(filter(None, [
            log.get('verified_last'),
            log.get('verified_first'),
            log.get('verified_sur'),
        ])) or '—'
        log['status_display'] = LOG_STATUS_LABELS.get(log.get('status'), log.get('status') or '—')

    # Список активных пользователей для формы
    users = _fetchall("""
        SELECT id, last_name, first_name, sur_name
        FROM users WHERE is_active = TRUE
        ORDER BY last_name, first_name
    """)
    for u in users:
        u['full_name'] = ' '.join(filter(None, [
            u.get('last_name'), u.get('first_name'), u.get('sur_name')
        ]))

    paginator = Paginator(logs, LOG_ITEMS_PER_PAGE)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    context = {
        'plan':               plan,
        'page_obj':           page_obj,
        'logs':               page_obj.object_list,
        'total_count':        len(logs),
        'can_edit':           can_edit,
        'log_status_choices': LOG_STATUS_CHOICES,
        'users':              users,
        'today':              today.isoformat(),
    }
    return render(request, 'core/maintenance_detail.html', context)


# ═════════════════════════════════════════════════════════════════
# Столбцы и экспорт для планов ТО ⭐ v3.31.0
# ═════════════════════════════════════════════════════════════════

@login_required
@require_POST
def save_maintenance_columns(request):
    """Сохранить выбранные столбцы для планов ТО."""
    try:
        data = json.loads(request.body)
        columns = data.get('columns', [])
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Некорректные данные'}, status=400)

    user = request.user
    prefs = user.ui_preferences or {}
    journal_columns = prefs.get('journal_columns', {})

    if columns == ['__reset__']:
        journal_columns.pop('MAINTENANCE', None)
    else:
        all_codes = {code for code, _ in MAINTENANCE_DISPLAYABLE_COLUMNS}
        valid = [c for c in columns if c in all_codes]
        if not valid:
            return JsonResponse({'error': 'Выберите хотя бы один столбец'}, status=400)
        journal_columns['MAINTENANCE'] = valid

    prefs['journal_columns'] = journal_columns
    user.ui_preferences = prefs
    user.save(update_fields=['ui_preferences'])
    return JsonResponse({'ok': True})


@login_required
@require_POST
def save_maintenance_column_widths(request):
    """Сохранить ширины столбцов для планов ТО."""
    try:
        data = json.loads(request.body)
        widths = data.get('widths', {})
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Некорректные данные'}, status=400)

    user = request.user
    prefs = user.ui_preferences or {}
    prefs['maintenance_column_widths'] = widths
    user.ui_preferences = prefs
    user.save(update_fields=['ui_preferences'])
    return JsonResponse({'ok': True})


@login_required
def export_maintenance_xlsx(request):
    """Экспорт планов ТО в XLSX."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not PermissionChecker.can_view(request.user, 'MAINTENANCE', 'access'):
        return HttpResponse('Нет доступа', status=403)

    # Те же фильтры, что и в maintenance_view
    search = request.GET.get('search', '').strip()
    lab_id = request.GET.get('lab_id', '')
    overdue_only = request.GET.get('overdue_only', '')
    sort = request.GET.get('sort', 'next_due_date')

    allowed_sorts = {
        'accounting_number', '-accounting_number',
        'equipment_type', '-equipment_type',
        'name', '-name',
        'next_due_date', '-next_due_date',
    }
    if sort not in allowed_sorts:
        sort = 'next_due_date'

    sort_col = sort.lstrip('-')
    sort_dir = 'DESC' if sort.startswith('-') else 'ASC'
    sort_col_map = {
        'accounting_number': 'e.accounting_number',
        'equipment_type': 'e.equipment_type',
        'name': 'emp.name',
        'next_due_date': 'emp.next_due_date',
    }
    order_by = f"{sort_col_map.get(sort_col, 'emp.next_due_date')} {sort_dir}"

    where_clauses, params = [], []
    if search:
        where_clauses.append("(e.accounting_number ILIKE %s OR emp.name ILIKE %s OR e.equipment_type ILIKE %s)")
        like = f'%{search}%'
        params += [like, like, like]
    if lab_id:
        where_clauses.append("e.laboratory_id = %s")
        params.append(int(lab_id))
    if overdue_only:
        where_clauses.append("emp.next_due_date < CURRENT_DATE")

    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    rows = _fetchall(f"""
        SELECT
            e.accounting_number, e.equipment_type,
            l.name AS laboratory_name, emp.name,
            emp.frequency_count, emp.frequency_period_value,
            emp.frequency_unit, emp.frequency_condition,
            emp.is_condition_based, emp.next_due_date, emp.notes
        FROM equipment_maintenance_plans emp
        JOIN equipment e ON e.id = emp.equipment_id
        LEFT JOIN laboratories l ON l.id = e.laboratory_id
        {where_sql}
        ORDER BY {order_by}
    """, params)

    for row in rows:
        row['frequency_display'] = _build_frequency_display(row)

    # Столбцы пользователя
    selected = _get_maintenance_user_columns(request.user)
    columns = [(code, MAINTENANCE_COLUMNS_DICT[code]) for code in selected if code in MAINTENANCE_COLUMNS_DICT]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Планы ТО'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(size=10)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'),
    )
    alt_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

    for col_idx, (code, name) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f'A1:{last_col}1'

    def _get_cell_value(row, code):
        if code == 'next_due_date':
            nd = row.get('next_due_date')
            return nd.strftime('%d.%m.%Y') if nd else ''
        return row.get(code, '') or ''

    for row_idx, row in enumerate(rows, 2):
        for col_idx, (code, _) in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_get_cell_value(row, code))
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = cell_alignment
        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill

    now_str = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
    filename = f'maintenance_plans_{now_str}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response