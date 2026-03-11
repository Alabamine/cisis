"""
equipment_views.py — Реестр оборудования
v3.29.0

Расположение: core/views/equipment_views.py

Новые маршруты в core/urls.py:
    path('workspace/equipment/', equipment_views.equipment_list, name='equipment_list'),
    path('workspace/equipment/<int:equipment_id>/', equipment_views.equipment_detail, name='equipment_detail'),
    path('workspace/equipment/<int:equipment_id>/edit/', equipment_views.equipment_edit, name='equipment_edit'),
    path('workspace/equipment/<int:equipment_id>/add-maintenance/', equipment_views.equipment_add_maintenance, name='equipment_add_maintenance'),
    path('workspace/equipment/save-columns/', equipment_views.save_equipment_columns, name='save_equipment_columns'),
    path('workspace/equipment/save-column-widths/', equipment_views.save_equipment_column_widths, name='save_equipment_column_widths'),
    path('workspace/equipment/filter-options/', equipment_views.equipment_filter_options, name='equipment_filter_options'),
    path('workspace/equipment/export/', equipment_views.export_equipment_xlsx, name='export_equipment_xlsx'),
"""

import json
from datetime import date, datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import connection
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.http import require_POST

from core.permissions import PermissionChecker
from core.models import (
    Equipment, EquipmentType, EquipmentStatus,
    Laboratory, User,
)
from core.models.base import AccreditationArea

ITEMS_PER_PAGE = 50
PER_PAGE_OPTIONS = [50, 100, 200]

# ═════════════════════════════════════════════════════════════════
# Определение столбцов для журнала оборудования
# ═════════════════════════════════════════════════════════════════

# (code, display_name) — все возможные столбцы
EQUIPMENT_DISPLAYABLE_COLUMNS = [
    ('accounting_number',       'Учётный номер'),
    ('equipment_type',          'Тип'),
    ('name',                    'Наименование'),
    ('inventory_number',        'Инвентарный номер'),
    ('laboratory',              'Подразделение'),
    ('status',                  'Статус'),
    ('responsible_person',      'Ответственный'),
    ('substitute_person',       'Замещающий'),
    ('accreditation_areas',     'Области аккредитации'),
    ('ownership',               'Принадлежность'),
    ('ownership_doc_number',    'Документ владения'),
    ('manufacturer',            'Производитель'),
    ('year_of_manufacture',     'Год выпуска'),
    ('factory_number',          'Заводской номер'),
    ('state_registry_number',   'Номер гос. реестра'),
    ('metrology_interval',      'Межповерочный интервал'),
    ('technical_documentation', 'Техническая документация'),
    ('intended_use',            'Назначение'),
    ('metrology_doc',           'Метрологические характеристики'),
    ('technical_specs',         'Технические характеристики'),
    ('software',                'Программное обеспечение'),
    ('operating_conditions',    'Условия эксплуатации'),
    ('commissioning_info',      'Ввод в эксплуатацию'),
    ('condition_on_receipt',    'Состояние при получении'),
    ('modifications',           'Модификации'),
    ('notes',                   'Примечания'),
]

EQUIPMENT_COLUMNS_DICT = {code: name for code, name in EQUIPMENT_DISPLAYABLE_COLUMNS}

# Столбцы по умолчанию
DEFAULT_EQUIPMENT_COLUMNS = [
    'accounting_number', 'equipment_type', 'name', 'inventory_number',
    'laboratory', 'status', 'responsible_person', 'manufacturer',
    'accreditation_areas',
]

# Фильтруемые столбцы
EQUIPMENT_FILTERABLE_COLUMNS = {
    'equipment_type': {
        'label': 'Тип оборудования',
        'type': 'select',
    },
    'status': {
        'label': 'Статус',
        'type': 'select',
    },
    'laboratory': {
        'label': 'Подразделение',
        'type': 'select',
    },
    'responsible_person': {
        'label': 'Ответственный',
        'type': 'select',
    },
    'accreditation_areas': {
        'label': 'Область аккредитации',
        'type': 'select',
    },
    'ownership': {
        'label': 'Принадлежность',
        'type': 'select',
    },
    'accounting_number': {
        'label': 'Учётный номер',
        'type': 'text',
    },
    'name': {
        'label': 'Наименование',
        'type': 'text',
    },
    'inventory_number': {
        'label': 'Инвентарный номер',
        'type': 'text',
    },
    'manufacturer': {
        'label': 'Производитель',
        'type': 'text',
    },
}


# ═════════════════════════════════════════════════════════════════
# Вспомогательные функции
# ═════════════════════════════════════════════════════════════════

def _get_user_selected_columns(user):
    """Возвращает список выбранных столбцов для пользователя."""
    prefs = user.ui_preferences or {}
    saved = prefs.get('journal_columns', {}).get('EQUIPMENT')
    if saved:
        all_codes = {code for code, _ in EQUIPMENT_DISPLAYABLE_COLUMNS}
        return [c for c in saved if c in all_codes]
    return list(DEFAULT_EQUIPMENT_COLUMNS)


def _build_base_queryset(user):
    """Строит базовый queryset оборудования."""
    qs = Equipment.objects.select_related(
        'laboratory', 'responsible_person', 'substitute_person'
    ).prefetch_related('accreditation_areas')
    return qs


def _apply_filters(queryset, params):
    """Применяет фильтры из GET-параметров."""

    # Select фильтры (множественный выбор)
    eq_type_values = params.getlist('equipment_type')
    if eq_type_values:
        queryset = queryset.filter(equipment_type__in=eq_type_values)

    status_values = params.getlist('status')
    if status_values:
        queryset = queryset.filter(status__in=status_values)

    lab_values = params.getlist('laboratory')
    if lab_values:
        queryset = queryset.filter(laboratory_id__in=lab_values)

    resp_values = params.getlist('responsible_person')
    if resp_values:
        queryset = queryset.filter(responsible_person_id__in=resp_values)

    area_values = params.getlist('accreditation_areas')
    if area_values:
        queryset = queryset.filter(accreditation_areas__id__in=area_values)

    ownership_values = params.getlist('ownership')
    if ownership_values:
        queryset = queryset.filter(ownership__in=ownership_values)

    # Текстовые фильтры
    acc_search = params.get('accounting_number_search', '').strip()
    if acc_search:
        queryset = queryset.filter(accounting_number__icontains=acc_search)

    name_search = params.get('name_search', '').strip()
    if name_search:
        queryset = queryset.filter(name__icontains=name_search)

    inv_search = params.get('inventory_number_search', '').strip()
    if inv_search:
        queryset = queryset.filter(inventory_number__icontains=inv_search)

    mfr_search = params.get('manufacturer_search', '').strip()
    if mfr_search:
        queryset = queryset.filter(manufacturer__icontains=mfr_search)

    return queryset


def _count_active_filters(params):
    """Подсчитывает количество активных фильтров."""
    count = 0
    filter_keys = [
        'equipment_type', 'status', 'laboratory',
        'responsible_person', 'accreditation_areas', 'ownership',
    ]
    for key in filter_keys:
        if params.getlist(key):
            count += 1
    for suffix in ('_search',):
        for key in params:
            if key.endswith(suffix) and params.get(key, '').strip():
                count += 1
    return count


def _get_filter_options(queryset):
    """Возвращает доступные варианты фильтров для queryset."""
    base_qs = queryset.order_by()

    options = {}

    # Тип оборудования
    existing_types = set(base_qs.values_list('equipment_type', flat=True).distinct())
    options['equipment_type'] = [
        {'value': t.value, 'label': t.label}
        for t in EquipmentType
        if t.value in existing_types
    ]

    # Статус
    existing_statuses = set(base_qs.values_list('status', flat=True).distinct())
    options['status'] = [
        {'value': s.value, 'label': s.label}
        for s in EquipmentStatus
        if s.value in existing_statuses
    ]

    # Подразделение
    labs = base_qs.values_list(
        'laboratory_id', 'laboratory__code_display', 'laboratory__name'
    ).distinct().order_by('laboratory__code_display')
    options['laboratory'] = [
        {'value': str(l[0]), 'label': f"{l[1]} — {l[2]}"}
        for l in labs if l[0]
    ]

    # Ответственный
    resp = base_qs.exclude(
        responsible_person__isnull=True
    ).values_list(
        'responsible_person_id',
        'responsible_person__last_name',
        'responsible_person__first_name',
    ).distinct().order_by('responsible_person__last_name')
    options['responsible_person'] = [
        {'value': str(r[0]), 'label': f"{r[1]} {r[2]}".strip()}
        for r in resp if r[0]
    ]

    # Области аккредитации
    areas = base_qs.values_list(
        'accreditation_areas__id', 'accreditation_areas__name'
    ).distinct().order_by('accreditation_areas__name')
    options['accreditation_areas'] = [
        {'value': str(a[0]), 'label': a[1]}
        for a in areas if a[0]
    ]

    # Принадлежность
    ownerships = base_qs.values_list('ownership', flat=True).distinct().order_by('ownership')
    options['ownership'] = [
        {'value': o, 'label': o}
        for o in ownerships if o
    ]

    return options


def _apply_sorting(queryset, sort_field, sort_dir):
    """Применяет сортировку."""
    if sort_field and sort_field in EQUIPMENT_COLUMNS_DICT:
        sort_map = {
            'laboratory': 'laboratory__code_display',
            'responsible_person': 'responsible_person__last_name',
            'substitute_person': 'substitute_person__last_name',
        }
        db_field = sort_map.get(sort_field, sort_field)
        if sort_dir == 'desc':
            db_field = f'-{db_field}'
        return queryset.order_by(db_field)
    return queryset.order_by('equipment_type', 'accounting_number')


def _get_export_value(eq, column_code):
    """Возвращает значение поля оборудования для экспорта в XLSX."""
    if column_code == 'laboratory':
        return eq.laboratory.code_display if eq.laboratory else ''
    elif column_code == 'responsible_person':
        return eq.responsible_person.full_name if eq.responsible_person else ''
    elif column_code == 'substitute_person':
        return eq.substitute_person.full_name if eq.substitute_person else ''
    elif column_code == 'equipment_type':
        return eq.get_equipment_type_display()
    elif column_code == 'status':
        return eq.get_status_display()
    elif column_code == 'accreditation_areas':
        areas = eq.accreditation_areas.all()
        return ', '.join(a.name for a in areas) if areas else ''
    else:
        val = getattr(eq, column_code, None)
        return val if val else ''


# ═════════════════════════════════════════════════════════════════
# Views
# ═════════════════════════════════════════════════════════════════

@login_required
def equipment_list(request):
    """Реестр оборудования: пагинация, фильтрация, кастомизация столбцов."""

    if not PermissionChecker.can_view(request.user, 'EQUIPMENT', 'access'):
        messages.error(request, 'У вас нет доступа к реестру оборудования')
        return redirect('workspace_home')

    user = request.user
    can_edit = PermissionChecker.can_edit(user, 'EQUIPMENT', 'access')

    # ─── Queryset ───
    equipment = _build_base_queryset(user)
    equipment = _apply_filters(equipment, request.GET)
    active_filter_count = _count_active_filters(request.GET)

    total_count = equipment.distinct().count()

    # ─── Статистика ───
    stats = {'total': total_count}
    base_qs = equipment.distinct()
    stats['operational'] = base_qs.filter(status='OPERATIONAL').count()
    stats['maintenance'] = base_qs.filter(status='MAINTENANCE').count()
    stats['calibration'] = base_qs.filter(status='CALIBRATION').count()

    # ─── Сортировка ───
    sort_field = request.GET.get('sort', '')
    sort_dir = request.GET.get('dir', 'asc')
    equipment = _apply_sorting(equipment, sort_field, sort_dir)

    equipment = equipment.distinct()

    # ─── Пагинация ───
    try:
        per_page = int(request.GET.get('per_page', ITEMS_PER_PAGE))
        if per_page not in PER_PAGE_OPTIONS:
            per_page = ITEMS_PER_PAGE
    except (ValueError, TypeError):
        per_page = ITEMS_PER_PAGE

    page_number = request.GET.get('page', 1)
    paginator = Paginator(equipment, per_page)
    page_obj = paginator.get_page(page_number)

    # ─── Столбцы ───
    all_codes = [code for code, _ in EQUIPMENT_DISPLAYABLE_COLUMNS]
    selected_columns = _get_user_selected_columns(user)

    visible_columns = [
        {'code': code, 'name': EQUIPMENT_COLUMNS_DICT[code]}
        for code in selected_columns
        if code in EQUIPMENT_COLUMNS_DICT
    ]

    all_available_columns = []
    # Сначала выбранные в их порядке
    for code in selected_columns:
        if code in EQUIPMENT_COLUMNS_DICT:
            all_available_columns.append({
                'code': code,
                'name': EQUIPMENT_COLUMNS_DICT[code],
                'selected': True,
            })
    # Затем невыбранные
    for code in all_codes:
        if code not in selected_columns and code in EQUIPMENT_COLUMNS_DICT:
            all_available_columns.append({
                'code': code,
                'name': EQUIPMENT_COLUMNS_DICT[code],
                'selected': False,
            })

    # ─── Фильтры ───
    filter_options = _get_filter_options(equipment)

    available_filters = {}
    for col_code, filter_config in EQUIPMENT_FILTERABLE_COLUMNS.items():
        available_filters[col_code] = {
            **filter_config,
            'options': filter_options.get(col_code, []),
        }

    current_filters = {}
    for key in EQUIPMENT_FILTERABLE_COLUMNS:
        values = request.GET.getlist(key)
        if values:
            current_filters[key] = values
    for suffix in ('_search',):
        for key in request.GET:
            if key.endswith(suffix):
                current_filters[key] = request.GET.get(key)

    # ─── URL params ───
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_string = query_params.urlencode()

    # ─── Ширины столбцов ───
    prefs = user.ui_preferences or {}
    column_widths = prefs.get('equipment_column_widths', {})

    return render(request, 'core/equipment_list.html', {
        'page_obj': page_obj,
        'equipment_items': page_obj.object_list,
        'visible_columns': visible_columns,
        'all_available_columns': all_available_columns,
        'available_filters': json.dumps(available_filters, ensure_ascii=False),
        'current_filters': json.dumps(current_filters, ensure_ascii=False),
        'active_filter_count': active_filter_count,
        'stats': stats,
        'user': user,
        'can_edit': can_edit,
        'query_string': query_string,
        'current_sort': sort_field,
        'current_dir': sort_dir,
        'total_count': total_count,
        'column_widths': json.dumps(column_widths),
        'per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
    })


@login_required
def equipment_detail(request, equipment_id):
    """Карточка оборудования."""
    if not PermissionChecker.can_view(request.user, 'EQUIPMENT', 'access'):
        messages.error(request, 'У вас нет доступа к реестру оборудования')
        return redirect('workspace_home')

    eq = get_object_or_404(Equipment, pk=equipment_id)
    can_edit = PermissionChecker.can_edit(request.user, 'EQUIPMENT', 'access')

    # Области аккредитации
    areas = eq.accreditation_areas.all().order_by('name')

    # История обслуживания (все записи)
    maintenance_history = eq.maintenance_history.select_related('performed_by').order_by('-maintenance_date')

    # Планы ТО
    maintenance_plans = eq.maintenance_plans.filter(is_active=True).order_by('name')

    # ⭐ v3.29.0: Статус поверки (последняя запись VERIFICATION с результатом SUITABLE)
    verification_status = None
    last_verification = maintenance_history.filter(
        maintenance_type='VERIFICATION',
        verification_result='SUITABLE',
        valid_until__isnull=False,
    ).first()

    if last_verification:
        from datetime import date as date_cls
        today = date_cls.today()
        days_left = (last_verification.valid_until - today).days
        if days_left < 0:
            verification_status = {
                'status': 'expired', 'label': 'Просрочена',
                'color': '#dc3545', 'bg': '#f8d7da',
                'days': abs(days_left),
                'record': last_verification,
            }
        elif days_left <= 30:
            verification_status = {
                'status': 'expiring', 'label': 'Истекает скоро',
                'color': '#856404', 'bg': '#fff3cd',
                'days': days_left,
                'record': last_verification,
            }
        elif days_left <= 90:
            verification_status = {
                'status': 'warning', 'label': f'Истекает через {days_left} дн.',
                'color': '#856404', 'bg': '#fff8e1',
                'days': days_left,
                'record': last_verification,
            }
        else:
            verification_status = {
                'status': 'valid', 'label': 'Действующая',
                'color': '#155724', 'bg': '#d4edda',
                'days': days_left,
                'record': last_verification,
            }

    # Последняя аттестация (для ИО)
    last_attestation = maintenance_history.filter(
        maintenance_type='ATTESTATION',
        valid_until__isnull=False,
    ).first()

    # Справочники для формы добавления
    from core.models.equipment import MaintenanceType
    maintenance_types = MaintenanceType.choices
    users_for_form = User.objects.filter(is_active=True).order_by('last_name', 'first_name')

    # ⭐ v3.30.0: Файлы оборудования
    can_upload_files = PermissionChecker.can_edit(request.user, 'FILES', 'equipment_files')
    can_delete_files = can_upload_files  # те же права

    context = {
        'eq': eq,
        'can_edit': can_edit,
        'areas': areas,
        'maintenance_history': maintenance_history[:20],
        'maintenance_plans': maintenance_plans,
        'verification_status': verification_status,
        'last_verification': last_verification,
        'last_attestation': last_attestation,
        'maintenance_types': maintenance_types,
        'users_for_form': users_for_form,
        # Файлы
        'can_upload_files': can_upload_files,
        'can_delete_files': can_delete_files,
    }
    return render(request, 'core/equipment_detail.html', context)


@login_required
@require_POST
def equipment_add_maintenance(request, equipment_id):
    """Добавить запись ТО/поверки из карточки оборудования."""
    if not PermissionChecker.can_edit(request.user, 'EQUIPMENT', 'access'):
        return JsonResponse({'error': 'Нет прав'}, status=403)

    eq = get_object_or_404(Equipment, pk=equipment_id)

    maintenance_type = request.POST.get('maintenance_type', '').strip()
    maintenance_date = request.POST.get('maintenance_date', '').strip()
    document_name = request.POST.get('document_name', '').strip()
    description = request.POST.get('description', '').strip()
    performed_by_id = request.POST.get('performed_by', '').strip()
    reason = request.POST.get('reason', '').strip()

    # Поверочные поля
    certificate_number = request.POST.get('certificate_number', '').strip()
    valid_until = request.POST.get('valid_until', '').strip()
    verification_organization = request.POST.get('verification_organization', '').strip()
    verification_result = request.POST.get('verification_result', '').strip()
    fgis_arshin_number = request.POST.get('fgis_arshin_number', '').strip()

    if not maintenance_type or not maintenance_date:
        messages.error(request, 'Тип и дата обязательны')
        return redirect('equipment_detail', equipment_id=equipment_id)

    with connection.cursor() as cur:
        cur.execute("""
            INSERT INTO equipment_maintenance
                (equipment_id, maintenance_type, maintenance_date, document_name,
                 description, performed_by_id, reason,
                 certificate_number, valid_until, verification_organization,
                 verification_result, fgis_arshin_number, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
        """, [
            eq.pk, maintenance_type, maintenance_date, document_name,
            description,
            int(performed_by_id) if performed_by_id else None, reason,
            certificate_number,
            valid_until if valid_until else None,
            verification_organization,
            verification_result,
            fgis_arshin_number,
        ])

    # Аудит
    try:
        from core.views.audit import log_action
        log_action(
            request, 'EQUIPMENT', eq.pk, 'MAINTENANCE_ADDED',
            extra_data={
                'equipment': f'{eq.accounting_number} — {eq.name}',
                'type': maintenance_type,
                'date': maintenance_date,
            }
        )
    except Exception:
        pass

    messages.success(request, f'Запись ТО добавлена для «{eq.name}»')
    return redirect('equipment_detail', equipment_id=equipment_id)


@login_required
def equipment_edit(request, equipment_id):
    """Редактирование оборудования."""
    if not PermissionChecker.can_edit(request.user, 'EQUIPMENT', 'access'):
        messages.error(request, 'У вас нет прав для редактирования оборудования')
        return redirect('equipment_detail', equipment_id=equipment_id)

    eq = get_object_or_404(Equipment, pk=equipment_id)

    laboratories = Laboratory.objects.filter(is_active=True).order_by('name')
    users = User.objects.filter(is_active=True).order_by('last_name', 'first_name')
    accreditation_areas = AccreditationArea.objects.filter(is_active=True).order_by('name')
    current_area_ids = set(eq.accreditation_areas.values_list('id', flat=True))

    if request.method == 'POST':
        errors = []

        accounting_number = request.POST.get('accounting_number', '').strip()
        equipment_type = request.POST.get('equipment_type', '').strip()
        name = request.POST.get('name', '').strip()
        inventory_number = request.POST.get('inventory_number', '').strip()
        lab_id = request.POST.get('laboratory', '').strip()
        status = request.POST.get('status', '').strip()

        if not accounting_number:
            errors.append('Учётный номер обязателен')
        if not name:
            errors.append('Наименование обязательно')
        if not lab_id:
            errors.append('Подразделение обязательно')

        if errors:
            for err in errors:
                messages.error(request, err)
        else:
            eq.accounting_number = accounting_number
            eq.equipment_type = equipment_type
            eq.name = name
            eq.inventory_number = inventory_number
            eq.laboratory_id = int(lab_id)
            eq.status = status

            # Ответственные
            resp_id = request.POST.get('responsible_person', '').strip()
            eq.responsible_person_id = int(resp_id) if resp_id else None
            sub_id = request.POST.get('substitute_person', '').strip()
            eq.substitute_person_id = int(sub_id) if sub_id else None

            # Производитель
            eq.manufacturer = request.POST.get('manufacturer', '').strip()
            year = request.POST.get('year_of_manufacture', '').strip()
            eq.year_of_manufacture = int(year) if year else None
            eq.factory_number = request.POST.get('factory_number', '').strip()
            eq.state_registry_number = request.POST.get('state_registry_number', '').strip()
            eq.ownership = request.POST.get('ownership', '').strip()
            eq.ownership_doc_number = request.POST.get('ownership_doc_number', '').strip()
            interval = request.POST.get('metrology_interval', '').strip()
            eq.metrology_interval = int(interval) if interval else None

            # Текстовые поля
            eq.technical_documentation = request.POST.get('technical_documentation', '').strip()
            eq.intended_use = request.POST.get('intended_use', '').strip()
            eq.metrology_doc = request.POST.get('metrology_doc', '').strip()
            eq.technical_specs = request.POST.get('technical_specs', '').strip()
            eq.software = request.POST.get('software', '').strip()
            eq.operating_conditions = request.POST.get('operating_conditions', '').strip()
            eq.commissioning_info = request.POST.get('commissioning_info', '').strip()
            eq.condition_on_receipt = request.POST.get('condition_on_receipt', '').strip()
            eq.modifications = request.POST.get('modifications', '').strip()
            eq.notes = request.POST.get('notes', '').strip()

            try:
                eq_before = Equipment.objects.get(pk=eq.pk)
                eq.save()

                # Обновляем области аккредитации (M2M)
                new_area_ids = set(
                    int(a) for a in request.POST.getlist('area_ids') if a.isdigit()
                )
                old_area_ids = current_area_ids

                with connection.cursor() as cur:
                    to_remove = old_area_ids - new_area_ids
                    if to_remove:
                        cur.execute(
                            "DELETE FROM equipment_accreditation_areas "
                            "WHERE equipment_id = %s AND accreditation_area_id = ANY(%s)",
                            [eq.pk, list(to_remove)]
                        )
                    to_add = new_area_ids - old_area_ids
                    for area_id in to_add:
                        cur.execute(
                            "INSERT INTO equipment_accreditation_areas "
                            "(equipment_id, accreditation_area_id) "
                            "VALUES (%s, %s) ON CONFLICT DO NOTHING",
                            [eq.pk, area_id]
                        )

                # Аудит
                try:
                    from core.views.audit import log_action
                    extra = {'equipment': f'{eq.accounting_number} — {eq.name}'}
                    TRACKED = [
                        ('name',               'Наименование'),
                        ('accounting_number',  'Учётный номер'),
                        ('factory_number',     'Заводской номер'),
                        ('inventory_number',   'Инвентарный номер'),
                        ('manufacturer',       'Производитель'),
                        ('equipment_type',     'Тип'),
                        ('status',             'Статус'),
                        ('metrology_interval', 'Межповерочный интервал'),
                        ('notes',              'Примечания'),
                    ]
                    for field, label in TRACKED:
                        old = str(getattr(eq_before, field) or '')
                        new = str(getattr(eq, field) or '')
                        if old != new:
                            log_action(
                                request, 'EQUIPMENT', eq.pk, 'EQUIPMENT_EDIT',
                                field_name=field,
                                old_value=old,
                                new_value=new,
                                extra_data=extra,
                            )
                except Exception:
                    pass

                messages.success(request, f'Оборудование «{eq.name}» обновлено')
                return redirect('equipment_detail', equipment_id=eq.pk)
            except Exception as e:
                messages.error(request, f'Ошибка сохранения: {e}')

    context = {
        'eq': eq,
        'laboratories': laboratories,
        'users': users,
        'equipment_types': EquipmentType.choices,
        'statuses': EquipmentStatus.choices,
        'accreditation_areas': accreditation_areas,
        'current_area_ids': current_area_ids,
    }
    return render(request, 'core/equipment_edit.html', context)


# ═════════════════════════════════════════════════════════════════
# AJAX endpoints
# ═════════════════════════════════════════════════════════════════

@login_required
@require_POST
def save_equipment_columns(request):
    """AJAX: сохранить выбранные столбцы для журнала оборудования."""
    try:
        data = json.loads(request.body)
        columns = data.get('columns', [])

        if not columns:
            return JsonResponse({'error': 'Список столбцов не может быть пустым'}, status=400)

        if columns == ['__reset__']:
            user = request.user
            prefs = user.ui_preferences or {}
            if 'journal_columns' in prefs and 'EQUIPMENT' in prefs['journal_columns']:
                del prefs['journal_columns']['EQUIPMENT']
                user.ui_preferences = prefs
                user.save(update_fields=['ui_preferences'])
            return JsonResponse({'status': 'ok', 'reset': True})

        all_codes = {code for code, _ in EQUIPMENT_DISPLAYABLE_COLUMNS}
        valid_columns = [c for c in columns if c in all_codes]

        if not valid_columns:
            return JsonResponse({'error': 'Ни один из столбцов не доступен'}, status=400)

        user = request.user
        prefs = user.ui_preferences or {}
        if 'journal_columns' not in prefs:
            prefs['journal_columns'] = {}
        prefs['journal_columns']['EQUIPMENT'] = valid_columns
        user.ui_preferences = prefs
        user.save(update_fields=['ui_preferences'])

        return JsonResponse({'status': 'ok', 'columns': valid_columns})

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Неверный формат JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def save_equipment_column_widths(request):
    """AJAX: сохранить ширины столбцов."""
    try:
        data = json.loads(request.body)
        widths = data.get('widths', {})

        user = request.user
        prefs = user.ui_preferences or {}
        prefs['equipment_column_widths'] = widths
        user.ui_preferences = prefs
        user.save(update_fields=['ui_preferences'])

        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def equipment_filter_options(request):
    """AJAX: каскадные фильтры."""
    if not PermissionChecker.can_view(request.user, 'EQUIPMENT', 'access'):
        return JsonResponse({'error': 'Нет доступа'}, status=403)

    equipment = _build_base_queryset(request.user)
    equipment = _apply_filters(equipment, request.GET).distinct()
    options = _get_filter_options(equipment)
    return JsonResponse(options)


@login_required
def export_equipment_xlsx(request):
    """Экспорт реестра оборудования в XLSX."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not PermissionChecker.can_view(request.user, 'EQUIPMENT', 'access'):
        return HttpResponse('Нет доступа', status=403)

    user = request.user

    equipment = _build_base_queryset(user)
    equipment = _apply_filters(equipment, request.GET).distinct()

    sort_field = request.GET.get('sort', '')
    sort_dir = request.GET.get('dir', 'asc')
    equipment = _apply_sorting(equipment, sort_field, sort_dir)

    selected_columns = _get_user_selected_columns(user)
    columns = [
        (code, EQUIPMENT_COLUMNS_DICT[code])
        for code in selected_columns
        if code in EQUIPMENT_COLUMNS_DICT
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Реестр оборудования'

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='4A90E2')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Arial', size=10)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    alt_fill = PatternFill('solid', fgColor='F8F9FA')

    for col_idx, (code, name) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.freeze_panes = 'A2'

    if columns:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f'A1:{last_col}1'

    row_idx = 2
    for eq in equipment:
        for col_idx, (code, name) in enumerate(columns, 1):
            value = _get_export_value(eq, code)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = cell_alignment

        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill

        row_idx += 1

    for col_idx, (code, name) in enumerate(columns, 1):
        max_len = len(name)
        for row in range(2, min(row_idx, 52)):
            val = ws.cell(row=row, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    now_str = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
    filename = f'equipment_{now_str}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ═════════════════════════════════════════════════════════════════
# ЖУРНАЛ ТО (все записи equipment_maintenance) ⭐ v3.30.0
# ═════════════════════════════════════════════════════════════════

MAINTENANCE_LOG_PER_PAGE = 50
MAINTENANCE_LOG_PER_PAGE_OPTIONS = [50, 100, 200]

# ═════════════════════════════════════════════════════════════════
# Столбцы для журнала поверок/аттестаций ⭐ v3.31.0
# ═════════════════════════════════════════════════════════════════

MAINTENANCE_LOG_DISPLAYABLE_COLUMNS = [
    ('accounting_number',       'Уч. номер'),
    ('equipment_name',          'Наименование'),
    ('laboratory',              'Подразделение'),
    ('maintenance_type',        'Тип'),
    ('maintenance_date',        'Дата'),
    ('document',                'Документ'),
    ('valid_until',             'Действ. до'),
    ('verification_result',     'Результат'),
    ('description',             'Описание'),
    ('performed_by',            'Выполнил'),
]

MAINTENANCE_LOG_COLUMNS_DICT = {code: name for code, name in MAINTENANCE_LOG_DISPLAYABLE_COLUMNS}

DEFAULT_MAINTENANCE_LOG_COLUMNS = [
    'accounting_number', 'equipment_name', 'laboratory',
    'maintenance_type', 'maintenance_date', 'document',
    'valid_until', 'verification_result', 'description', 'performed_by',
]


def _get_maintenance_log_user_columns(user):
    """Возвращает выбранные столбцы для журнала поверок."""
    prefs = user.ui_preferences or {}
    saved = prefs.get('journal_columns', {}).get('MAINTENANCE_LOG')
    if saved:
        all_codes = {code for code, _ in MAINTENANCE_LOG_DISPLAYABLE_COLUMNS}
        return [c for c in saved if c in all_codes]
    return list(DEFAULT_MAINTENANCE_LOG_COLUMNS)

MAINTENANCE_TYPE_LABELS = {
    'VERIFICATION': 'Поверка',
    'ATTESTATION': 'Аттестация',
    'REPAIR': 'Ремонт',
    'MODIFICATION': 'Модификация',
    'CALIBRATION': 'Калибровка',
    'CONSERVATION': 'Консервация',
}

VERIFICATION_RESULT_LABELS = {
    'SUITABLE': 'Пригоден',
    'UNSUITABLE': 'Непригоден',
}


@login_required
def equipment_maintenance_log(request):
    """
    Журнал ТО — все записи equipment_maintenance по всему оборудованию.
    Доступ через права EQUIPMENT (тот же журнал, что и реестр).
    """
    if not PermissionChecker.can_view(request.user, 'EQUIPMENT', 'access'):
        messages.error(request, 'У вас нет доступа к реестру оборудования')
        return redirect('workspace_home')

    can_edit = PermissionChecker.can_edit(request.user, 'EQUIPMENT', 'access')

    # ─── Фильтры ───
    f_type = request.GET.getlist('maintenance_type')
    f_lab = request.GET.getlist('laboratory')
    f_result = request.GET.getlist('verification_result')
    f_search = request.GET.get('search', '').strip()
    f_date_from = request.GET.get('date_from', '').strip()
    f_date_to = request.GET.get('date_to', '').strip()

    # ─── Queryset ───
    from core.models.equipment import EquipmentMaintenance
    qs = EquipmentMaintenance.objects.select_related(
        'equipment', 'equipment__laboratory', 'performed_by'
    )

    if f_type:
        qs = qs.filter(maintenance_type__in=f_type)
    if f_lab:
        qs = qs.filter(equipment__laboratory_id__in=f_lab)
    if f_result:
        qs = qs.filter(verification_result__in=f_result)
    if f_search:
        qs = qs.filter(
            Q(equipment__accounting_number__icontains=f_search) |
            Q(equipment__name__icontains=f_search) |
            Q(document_name__icontains=f_search) |
            Q(certificate_number__icontains=f_search)
        )
    if f_date_from:
        qs = qs.filter(maintenance_date__gte=f_date_from)
    if f_date_to:
        qs = qs.filter(maintenance_date__lte=f_date_to)

    # ─── Подсчёт активных фильтров ───
    active_filter_count = 0
    if f_type: active_filter_count += 1
    if f_lab: active_filter_count += 1
    if f_result: active_filter_count += 1
    if f_search: active_filter_count += 1
    if f_date_from or f_date_to: active_filter_count += 1

    total_count = qs.count()

    # ─── Сортировка ───
    sort_field = request.GET.get('sort', 'maintenance_date')
    sort_dir = request.GET.get('dir', 'desc')

    sort_map = {
        'maintenance_date': 'maintenance_date',
        'maintenance_type': 'maintenance_type',
        'accounting_number': 'equipment__accounting_number',
        'equipment_name': 'equipment__name',
        'laboratory': 'equipment__laboratory__code_display',
        'valid_until': 'valid_until',
        'performed_by': 'performed_by__last_name',
    }
    db_sort = sort_map.get(sort_field, 'maintenance_date')
    if sort_dir == 'desc':
        db_sort = f'-{db_sort}'
    qs = qs.order_by(db_sort)

    # ─── Пагинация ───
    try:
        per_page = int(request.GET.get('per_page', MAINTENANCE_LOG_PER_PAGE))
        if per_page not in MAINTENANCE_LOG_PER_PAGE_OPTIONS:
            per_page = MAINTENANCE_LOG_PER_PAGE
    except (ValueError, TypeError):
        per_page = MAINTENANCE_LOG_PER_PAGE

    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    # ─── Расчёт valid_until если не заполнено ───
    from datetime import date
    import calendar
    records = list(page_obj.object_list)
    for rec in records:
        if not rec.valid_until and rec.maintenance_date and rec.equipment.metrology_interval:
            d = rec.maintenance_date
            month = d.month - 1 + rec.equipment.metrology_interval
            year = d.year + month // 12
            month = month % 12 + 1
            day = min(d.day, calendar.monthrange(year, month)[1])
            rec.valid_until = date(year, month, day)
 

    # ─── Статистика ───
    stats = {
        'total': total_count,
        'verification': qs.filter(maintenance_type='VERIFICATION').count(),
        'attestation': qs.filter(maintenance_type='ATTESTATION').count(),
        'repair': qs.filter(maintenance_type='REPAIR').count(),
        'modification': qs.filter(maintenance_type='MODIFICATION').count(),
        'calibration': qs.filter(maintenance_type='CALIBRATION').count(),
        'conservation': qs.filter(maintenance_type='CONSERVATION').count(),     
    }

    # ─── Справочники для фильтров ───
    laboratories = Laboratory.objects.filter(
        is_active=True, department_type='LAB'
    ).order_by('code_display')

    # ─── URL params ───
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_string = query_params.urlencode()

    # ─── Столбцы ⭐ v3.31.0 ───
    selected_columns = _get_maintenance_log_user_columns(request.user)
    visible_columns = [
        {'code': code, 'name': MAINTENANCE_LOG_COLUMNS_DICT[code]}
        for code in selected_columns
        if code in MAINTENANCE_LOG_COLUMNS_DICT
    ]
    all_available_columns = []
    for code in selected_columns:
        if code in MAINTENANCE_LOG_COLUMNS_DICT:
            all_available_columns.append({'code': code, 'name': MAINTENANCE_LOG_COLUMNS_DICT[code], 'selected': True})
    for code, _ in MAINTENANCE_LOG_DISPLAYABLE_COLUMNS:
        if code not in selected_columns:
            all_available_columns.append({'code': code, 'name': MAINTENANCE_LOG_COLUMNS_DICT[code], 'selected': False})

    prefs = request.user.ui_preferences or {}
    column_widths = prefs.get('maintenance_log_column_widths', {})

    context = {
        'page_obj': page_obj,
        'records': records,
        'total_count': total_count,
        'stats': stats,
        'can_edit': can_edit,
        'laboratories': laboratories,
        'active_filter_count': active_filter_count,
        'query_string': query_string,
        'current_sort': sort_field,
        'current_dir': sort_dir,
        # Текущие фильтры
        'f_type': f_type,
        'f_lab': f_lab,
        'f_result': f_result,
        'f_search': f_search,
        'f_date_from': f_date_from,
        'f_date_to': f_date_to,
        # Справочники
        'maintenance_type_choices': [
            ('VERIFICATION', 'Поверка'),
            ('ATTESTATION', 'Аттестация'),
            ('REPAIR', 'Ремонт'),
            ('MODIFICATION',  'Модификация'),
            ('CALIBRATION',   'Калибровка'),
            ('CONSERVATION',  'Консервация'),
        ],
        'verification_result_choices': [
            ('SUITABLE', 'Пригоден'),
            ('UNSUITABLE', 'Непригоден'),
        ],
        # Флаг для табов
        'active_tab': 'maintenance_log',
        'per_page': per_page,
        'per_page_options': MAINTENANCE_LOG_PER_PAGE_OPTIONS,
        'user': request.user,
        'visible_columns': visible_columns,
        'all_available_columns': all_available_columns,
        'column_widths': json.dumps(column_widths),
    }
    return render(request, 'core/equipment_maintenance_log.html', context)


@login_required
def export_maintenance_log_xlsx(request):
    """Экспорт журнала ТО в XLSX."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not PermissionChecker.can_view(request.user, 'EQUIPMENT', 'access'):
        return HttpResponse('Нет доступа', status=403)

    from core.models.equipment import EquipmentMaintenance
    qs = EquipmentMaintenance.objects.select_related(
        'equipment', 'equipment__laboratory', 'performed_by'
    )

    # Применяем те же фильтры
    f_type = request.GET.getlist('maintenance_type')
    f_lab = request.GET.getlist('laboratory')
    f_result = request.GET.getlist('verification_result')
    f_search = request.GET.get('search', '').strip()
    f_date_from = request.GET.get('date_from', '').strip()
    f_date_to = request.GET.get('date_to', '').strip()

    if f_type:
        qs = qs.filter(maintenance_type__in=f_type)
    if f_lab:
        qs = qs.filter(equipment__laboratory_id__in=f_lab)
    if f_result:
        qs = qs.filter(verification_result__in=f_result)
    if f_search:
        qs = qs.filter(
            Q(equipment__accounting_number__icontains=f_search) |
            Q(equipment__name__icontains=f_search) |
            Q(document_name__icontains=f_search) |
            Q(certificate_number__icontains=f_search)
        )
    if f_date_from:
        qs = qs.filter(maintenance_date__gte=f_date_from)
    if f_date_to:
        qs = qs.filter(maintenance_date__lte=f_date_to)

    sort_field = request.GET.get('sort', 'maintenance_date')
    sort_dir = request.GET.get('dir', 'desc')
    sort_map = {
        'maintenance_date': 'maintenance_date',
        'maintenance_type': 'maintenance_type',
        'accounting_number': 'equipment__accounting_number',
        'equipment_name': 'equipment__name',
        'laboratory': 'equipment__laboratory__code_display',
        'valid_until': 'valid_until',
        'performed_by': 'performed_by__last_name',
    }
    db_sort = sort_map.get(sort_field, 'maintenance_date')
    if sort_dir == 'desc':
        db_sort = f'-{db_sort}'
    qs = qs.order_by(db_sort)

    # Создаём XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = 'Журнал ТО'

    columns = [
        ('Учётный номер', 18),
        ('Наименование', 30),
        ('Подразделение', 12),
        ('Тип ТО', 14),
        ('Дата', 12),
        ('Документ', 20),
        ('Свидетельство', 22),
        ('Действительно до', 16),
        ('Результат', 14),
        ('Организация', 25),
        ('Аршин', 18),
        ('Описание', 30),
        ('Выполнил', 20),
    ]

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='4A90E2')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Arial', size=10)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    alt_fill = PatternFill('solid', fgColor='F8F9FA')

    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f'A1:{last_col}1'

    row_idx = 2
    for rec in qs:
        values = [
            rec.equipment.accounting_number if rec.equipment else '',
            rec.equipment.name if rec.equipment else '',
            rec.equipment.laboratory.code_display if rec.equipment and rec.equipment.laboratory else '',
            MAINTENANCE_TYPE_LABELS.get(rec.maintenance_type, rec.maintenance_type),
            rec.maintenance_date.strftime('%d.%m.%Y') if rec.maintenance_date else '',
            rec.document_name or '',
            rec.certificate_number or '',
            rec.valid_until.strftime('%d.%m.%Y') if rec.valid_until else '',
            VERIFICATION_RESULT_LABELS.get(rec.verification_result, rec.verification_result or ''),
            rec.verification_organization or '',
            rec.fgis_arshin_number or '',
            rec.description or '',
            rec.performed_by.full_name if rec.performed_by else '',
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = cell_alignment

        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill

        row_idx += 1

    now_str = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
    filename = f'maintenance_log_{now_str}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ═════════════════════════════════════════════════════════════════
# СОЗДАНИЕ ПЛАНА ТО С КАРТОЧКИ ОБОРУДОВАНИЯ ⭐ v3.30.0
# ═════════════════════════════════════════════════════════════════

@login_required
@require_POST
def equipment_add_plan(request, equipment_id):
    """Добавить план ТО из карточки оборудования."""
    if not PermissionChecker.can_edit(request.user, 'EQUIPMENT', 'access'):
        return JsonResponse({'error': 'Нет прав'}, status=403)

    eq = get_object_or_404(Equipment, pk=equipment_id)

    plan_name = request.POST.get('plan_name', '').strip()
    frequency_count = request.POST.get('frequency_count', '').strip()
    frequency_period_value = request.POST.get('frequency_period_value', '').strip()
    frequency_unit = request.POST.get('frequency_unit', '').strip()
    if not frequency_count or not frequency_period_value or not frequency_unit:
        frequency_count = ''
        frequency_period_value = ''
        frequency_unit = ''
        # Без календаря — автоматически «по условию»
        if not is_condition_based:
            is_condition_based = True
    frequency_condition = request.POST.get('frequency_condition', '').strip()
    is_condition_based = request.POST.get('is_condition_based') == 'on'
    next_due_date = request.POST.get('next_due_date', '').strip()
    plan_notes = request.POST.get('plan_notes', '').strip()

    if not plan_name:
        messages.error(request, 'Название плана обязательно')
        return redirect('equipment_detail', equipment_id=equipment_id)

    with connection.cursor() as cur:
        cur.execute("""
            INSERT INTO equipment_maintenance_plans
                (equipment_id, name, frequency_count, frequency_period_value,
                 frequency_unit, frequency_condition, is_condition_based,
                 next_due_date, is_active, notes, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, TRUE, %s, NOW(), NOW())
        """, [
            eq.pk,
            plan_name,
            int(frequency_count) if frequency_count else None,
            int(frequency_period_value) if frequency_period_value else None,
            frequency_unit if frequency_unit else None,
            frequency_condition,
            is_condition_based,
            next_due_date if next_due_date else None,
            plan_notes,
        ])

    # Аудит
    try:
        from core.views.audit import log_action
        log_action(
            request, 'EQUIPMENT', eq.pk, 'PLAN_ADDED',
            extra_data={
                'equipment': f'{eq.accounting_number} — {eq.name}',
                'plan': plan_name,
            }
        )
    except Exception:
        pass

    messages.success(request, f'План ТО «{plan_name}» добавлен для «{eq.name}»')
    return redirect('equipment_detail', equipment_id=equipment_id)

# ═════════════════════════════════════════════════════════════════
# РЕДАКТИРОВАНИЕ ПЛАНА ТО ⭐ v3.30.0
# ═════════════════════════════════════════════════════════════════

@login_required
@require_POST
def equipment_edit_plan(request, equipment_id, plan_id):
    """Редактировать план ТО из карточки оборудования."""
    if not PermissionChecker.can_edit(request.user, 'EQUIPMENT', 'access'):
        return JsonResponse({'error': 'Нет прав'}, status=403)

    eq = get_object_or_404(Equipment, pk=equipment_id)

    plan_name = request.POST.get('plan_name', '').strip()
    frequency_count = request.POST.get('frequency_count', '').strip()
    frequency_period_value = request.POST.get('frequency_period_value', '').strip()
    frequency_unit = request.POST.get('frequency_unit', '').strip()

    if not frequency_count or not frequency_period_value or not frequency_unit:
        frequency_count = ''
        frequency_period_value = ''
        frequency_unit = ''

    frequency_condition = request.POST.get('frequency_condition', '').strip()
    is_condition_based = request.POST.get('is_condition_based') == 'on'

    if not frequency_count and not is_condition_based:
        is_condition_based = True

    next_due_date = request.POST.get('next_due_date', '').strip()
    plan_notes = request.POST.get('plan_notes', '').strip()
    is_active = request.POST.get('is_active') != 'off'

    if not plan_name:
        messages.error(request, 'Название плана обязательно')
        return redirect('equipment_detail', equipment_id=equipment_id)

    with connection.cursor() as cur:
        cur.execute("""
            UPDATE equipment_maintenance_plans
            SET name = %s, frequency_count = %s, frequency_period_value = %s,
                frequency_unit = %s, frequency_condition = %s, is_condition_based = %s,
                next_due_date = %s, is_active = %s, notes = %s, updated_at = NOW()
            WHERE id = %s AND equipment_id = %s
        """, [
            plan_name,
            int(frequency_count) if frequency_count else None,
            int(frequency_period_value) if frequency_period_value else None,
            frequency_unit if frequency_unit else None,
            frequency_condition,
            is_condition_based,
            next_due_date if next_due_date else None,
            is_active,
            plan_notes,
            plan_id, eq.pk,
        ])

    try:
        from core.views.audit import log_action
        log_action(
            request, 'EQUIPMENT', eq.pk, 'PLAN_EDITED',
            extra_data={
                'equipment': f'{eq.accounting_number} — {eq.name}',
                'plan': plan_name,
            }
        )
    except Exception:
        pass

    messages.success(request, f'План ТО «{plan_name}» обновлён')
    return redirect('equipment_detail', equipment_id=equipment_id)


@login_required
@require_POST
def equipment_delete_plan(request, equipment_id, plan_id):
    """Деактивировать план ТО."""
    if not PermissionChecker.can_edit(request.user, 'EQUIPMENT', 'access'):
        return JsonResponse({'error': 'Нет прав'}, status=403)

    eq = get_object_or_404(Equipment, pk=equipment_id)

    with connection.cursor() as cur:
        cur.execute("""
            UPDATE equipment_maintenance_plans
            SET is_active = FALSE, updated_at = NOW()
            WHERE id = %s AND equipment_id = %s
        """, [plan_id, eq.pk])

    try:
        from core.views.audit import log_action
        log_action(
            request, 'EQUIPMENT', eq.pk, 'PLAN_DEACTIVATED',
            extra_data={'equipment': f'{eq.accounting_number} — {eq.name}'}
        )
    except Exception:
        pass

    messages.success(request, 'План ТО деактивирован')
    return redirect('equipment_detail', equipment_id=equipment_id)


# ═════════════════════════════════════════════════════════════════
# Столбцы для журнала поверок/аттестаций ⭐ v3.31.0
# ═════════════════════════════════════════════════════════════════

@login_required
@require_POST
def save_maintenance_log_columns(request):
    """Сохранить выбранные столбцы для журнала поверок."""
    try:
        data = json.loads(request.body)
        columns = data.get('columns', [])
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Некорректные данные'}, status=400)

    user = request.user
    prefs = user.ui_preferences or {}
    journal_columns = prefs.get('journal_columns', {})

    if columns == ['__reset__']:
        journal_columns.pop('MAINTENANCE_LOG', None)
    else:
        all_codes = {code for code, _ in MAINTENANCE_LOG_DISPLAYABLE_COLUMNS}
        valid = [c for c in columns if c in all_codes]
        if not valid:
            return JsonResponse({'error': 'Выберите хотя бы один столбец'}, status=400)
        journal_columns['MAINTENANCE_LOG'] = valid

    prefs['journal_columns'] = journal_columns
    user.ui_preferences = prefs
    user.save(update_fields=['ui_preferences'])
    return JsonResponse({'ok': True})


@login_required
@require_POST
def save_maintenance_log_column_widths(request):
    """Сохранить ширины столбцов для журнала поверок."""
    try:
        data = json.loads(request.body)
        widths = data.get('widths', {})
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Некорректные данные'}, status=400)

    user = request.user
    prefs = user.ui_preferences or {}
    prefs['maintenance_log_column_widths'] = widths
    user.ui_preferences = prefs
    user.save(update_fields=['ui_preferences'])
    return JsonResponse({'ok': True})