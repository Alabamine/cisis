"""
file_manager_views.py — Файловый менеджер
v3.31.0

Расположение: core/views/file_manager_views.py

Маршруты в core/urls.py:
    path('workspace/files/', file_manager_views.file_manager, name='file_manager'),
    path('workspace/files/export/', file_manager_views.export_files_xlsx, name='export_files_xlsx'),
"""

import json
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST

from core.models import File, Laboratory
from core.models.files import FileCategory, FileType
from core.permissions import PermissionChecker

FILES_PER_PAGE = 50
PER_PAGE_OPTIONS = [50, 100, 200]

# ═════════════════════════════════════════════════════════════════
# Метки типов файлов для отображения
# ═════════════════════════════════════════════════════════════════

FILE_TYPE_LABELS = {}
for _cat, _choices in FileType.CHOICES_BY_CATEGORY.items():
    for _val, _label in _choices:
        FILE_TYPE_LABELS[_val] = _label

CATEGORY_LABELS = dict(FileCategory.CHOICES)

# Доступные категории (расширяемый список)
AVAILABLE_CATEGORIES = [
    ('EQUIPMENT', '🔬 Оборудование'),
    # ('SAMPLE', '🧪 Образцы'),      # TODO: v3.32+
    # ('CLIENT', '👥 Клиенты'),       # TODO: v3.32+
    # ('STANDARD', '📖 Стандарты'),   # TODO: v3.32+
    # ('QMS', '📋 СМК'),             # TODO: v3.32+
]
AVAILABLE_CATEGORY_CODES = [c[0] for c in AVAILABLE_CATEGORIES]

# ═════════════════════════════════════════════════════════════════
# Столбцы файлового менеджера (EQUIPMENT)
# ═════════════════════════════════════════════════════════════════

FM_EQUIPMENT_COLUMNS = [
    ('accounting_number', 'Уч. номер'),
    ('equipment_name',    'Оборудование'),
    ('laboratory',        'Подразделение'),
    ('file_type',         'Тип файла'),
    ('original_name',     'Файл'),
    ('file_size',         'Размер'),
    ('uploaded_by',       'Загрузил'),
    ('uploaded_at',       'Дата'),
    ('download',          ''),
]

FM_EQUIPMENT_COLUMNS_DICT = {code: name for code, name in FM_EQUIPMENT_COLUMNS}

DEFAULT_FM_EQUIPMENT_COLUMNS = [
    'accounting_number', 'equipment_name', 'laboratory',
    'file_type', 'original_name', 'file_size',
    'uploaded_by', 'uploaded_at', 'download',
]


def _get_fm_user_columns(user, category):
    """Возвращает выбранные столбцы для файлового менеджера."""
    prefs = user.ui_preferences or {}
    key = f'FM_{category}'
    saved = prefs.get('journal_columns', {}).get(key)
    if saved:
        all_codes = {code for code, _ in FM_EQUIPMENT_COLUMNS}
        return [c for c in saved if c in all_codes]
    return list(DEFAULT_FM_EQUIPMENT_COLUMNS)


# ═════════════════════════════════════════════════════════════════
# Главная страница файлового менеджера
# ═════════════════════════════════════════════════════════════════

@login_required
def file_manager(request):
    """Файловый менеджер — единая страница просмотра всех файлов."""

    if not PermissionChecker.can_view(request.user, 'FILES', 'equipment_files'):
        messages.error(request, 'У вас нет доступа к файловому менеджеру')
        return redirect('workspace_home')

    # ─── Текущая категория (из GET или дефолт) ───
    current_category = request.GET.get('category', 'EQUIPMENT')
    if current_category not in AVAILABLE_CATEGORY_CODES:
        current_category = 'EQUIPMENT'

    # ─── Queryset ───
    qs = File.objects.filter(
        category=current_category,
        is_deleted=False,
        current_version=True,
    ).select_related(
        'equipment', 'equipment__laboratory',
        'uploaded_by',
    ).order_by('-uploaded_at')

    # ─── Фильтры ───
    f_type = request.GET.getlist('file_type')
    f_lab = request.GET.getlist('laboratory')
    f_search = request.GET.get('search', '').strip()

    if current_category == 'EQUIPMENT':
        if f_type:
            qs = qs.filter(file_type__in=f_type)
        if f_lab:
            qs = qs.filter(equipment__laboratory_id__in=f_lab)
        if f_search:
            qs = qs.filter(
                Q(equipment__accounting_number__icontains=f_search) |
                Q(equipment__name__icontains=f_search) |
                Q(original_name__icontains=f_search) |
                Q(description__icontains=f_search)
            )

    # ─── Подсчёт фильтров ───
    active_filter_count = 0
    if f_type: active_filter_count += 1
    if f_lab: active_filter_count += 1
    if f_search: active_filter_count += 1

    total_count = qs.count()

    # ─── Сортировка ───
    sort_field = request.GET.get('sort', 'uploaded_at')
    sort_dir = request.GET.get('dir', 'desc')

    sort_map = {
        'uploaded_at': 'uploaded_at',
        'original_name': 'original_name',
        'file_type': 'file_type',
        'file_size': 'file_size',
        'equipment_name': 'equipment__name',
        'accounting_number': 'equipment__accounting_number',
        'laboratory': 'equipment__laboratory__code_display',
        'uploaded_by': 'uploaded_by__last_name',
    }
    db_sort = sort_map.get(sort_field, 'uploaded_at')
    if sort_dir == 'desc':
        db_sort = f'-{db_sort}'
    qs = qs.order_by(db_sort)

    # ─── Пагинация ───
    try:
        per_page = int(request.GET.get('per_page', FILES_PER_PAGE))
        if per_page not in PER_PAGE_OPTIONS:
            per_page = FILES_PER_PAGE
    except (ValueError, TypeError):
        per_page = FILES_PER_PAGE

    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    # Аннотация: человекочитаемые метки типов
    for f in page_obj.object_list:
        f.type_label = FILE_TYPE_LABELS.get(f.file_type, f.file_type or '—')

    # ─── Статистика по типам ───
    stats = {}
    if current_category == 'EQUIPMENT':
        type_choices = FileType.CHOICES_BY_CATEGORY.get(FileCategory.EQUIPMENT, [])
        for val, label in type_choices:
            cnt = qs.filter(file_type=val).count()
            if cnt:
                stats[val] = {'label': label, 'count': cnt}
        stats['_total'] = total_count

    # ─── Справочники для фильтров ───
    laboratories = Laboratory.objects.filter(
        is_active=True, department_type='LAB'
    ).order_by('code_display')

    file_type_choices = FileType.CHOICES_BY_CATEGORY.get(current_category, [])

    # ─── Столбцы ───
    selected_columns = _get_fm_user_columns(request.user, current_category)
    visible_columns = [
        {'code': code, 'name': FM_EQUIPMENT_COLUMNS_DICT[code]}
        for code in selected_columns
        if code in FM_EQUIPMENT_COLUMNS_DICT
    ]
    all_available_columns = []
    for code in selected_columns:
        if code in FM_EQUIPMENT_COLUMNS_DICT:
            all_available_columns.append({'code': code, 'name': FM_EQUIPMENT_COLUMNS_DICT[code], 'selected': True})
    for code, _ in FM_EQUIPMENT_COLUMNS:
        if code not in selected_columns:
            all_available_columns.append({'code': code, 'name': FM_EQUIPMENT_COLUMNS_DICT[code], 'selected': False})

    prefs = request.user.ui_preferences or {}
    column_widths = prefs.get('fm_column_widths', {}).get(current_category, {})

    # ─── URL params (без page) ───
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_string = query_params.urlencode()

    context = {
        'page_obj': page_obj,
        'files': page_obj.object_list,
        'total_count': total_count,
        'stats': stats,
        'user': request.user,
        'active_filter_count': active_filter_count,
        'query_string': query_string,
        'current_sort': sort_field,
        'current_dir': sort_dir,
        'per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
        # Категории
        'available_categories': AVAILABLE_CATEGORIES,
        'current_category': current_category,
        'current_category_label': dict(AVAILABLE_CATEGORIES).get(current_category, current_category),
        # Фильтры
        'f_type': f_type,
        'f_lab': f_lab,
        'f_search': f_search,
        'file_type_choices': file_type_choices,
        'laboratories': laboratories,
        'file_type_labels': FILE_TYPE_LABELS,
        'visible_columns': visible_columns,
        'all_available_columns': all_available_columns,
        'column_widths': json.dumps(column_widths),
    }
    return render(request, 'core/file_manager.html', context)


# ═════════════════════════════════════════════════════════════════
# Экспорт в XLSX
# ═════════════════════════════════════════════════════════════════

@login_required
def export_files_xlsx(request):
    """Экспорт файлов текущей категории в XLSX."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not PermissionChecker.can_view(request.user, 'FILES', 'equipment_files'):
        return HttpResponse('Нет доступа', status=403)

    current_category = request.GET.get('category', 'EQUIPMENT')

    qs = File.objects.filter(
        category=current_category,
        is_deleted=False,
        current_version=True,
    ).select_related(
        'equipment', 'equipment__laboratory', 'uploaded_by',
    )

    # Применяем те же фильтры
    f_type = request.GET.getlist('file_type')
    f_lab = request.GET.getlist('laboratory')
    f_search = request.GET.get('search', '').strip()

    if current_category == 'EQUIPMENT':
        if f_type:
            qs = qs.filter(file_type__in=f_type)
        if f_lab:
            qs = qs.filter(equipment__laboratory_id__in=f_lab)
        if f_search:
            qs = qs.filter(
                Q(equipment__accounting_number__icontains=f_search) |
                Q(equipment__name__icontains=f_search) |
                Q(original_name__icontains=f_search)
            )

    qs = qs.order_by('-uploaded_at')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Файлы'

    columns = [
        ('Уч. номер', 14),
        ('Оборудование', 30),
        ('Подразделение', 12),
        ('Тип файла', 22),
        ('Имя файла', 35),
        ('Размер', 12),
        ('Загрузил', 20),
        ('Дата', 12),
        ('Описание', 30),
    ]

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(size=10)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'),
    )
    alt_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f'A1:{last_col}1'

    for row_idx, f in enumerate(qs[:5000], 2):  # лимит 5000
        values = [
            f.equipment.accounting_number if f.equipment else '',
            f.equipment.name if f.equipment else '',
            f.equipment.laboratory.code_display if f.equipment and f.equipment.laboratory else '',
            FILE_TYPE_LABELS.get(f.file_type, f.file_type),
            f.original_name,
            f.size_display,
            f.uploaded_by.full_name if f.uploaded_by else '',
            f.uploaded_at.strftime('%d.%m.%Y') if f.uploaded_at else '',
            f.description or '',
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = cell_alignment
        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill

    now_str = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
    filename = f'files_{current_category.lower()}_{now_str}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ═════════════════════════════════════════════════════════════════
# Сохранение столбцов и ширин
# ═════════════════════════════════════════════════════════════════

@login_required
@require_POST
def save_fm_columns(request):
    """Сохранить выбранные столбцы для файлового менеджера."""
    try:
        data = json.loads(request.body)
        columns = data.get('columns', [])
        category = data.get('category', 'EQUIPMENT')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Некорректные данные'}, status=400)

    user = request.user
    prefs = user.ui_preferences or {}
    journal_columns = prefs.get('journal_columns', {})
    key = f'FM_{category}'

    if columns == ['__reset__']:
        journal_columns.pop(key, None)
    else:
        all_codes = {code for code, _ in FM_EQUIPMENT_COLUMNS}
        valid = [c for c in columns if c in all_codes]
        if not valid:
            return JsonResponse({'error': 'Выберите хотя бы один столбец'}, status=400)
        journal_columns[key] = valid

    prefs['journal_columns'] = journal_columns
    user.ui_preferences = prefs
    user.save(update_fields=['ui_preferences'])
    return JsonResponse({'ok': True})


@login_required
@require_POST
def save_fm_column_widths(request):
    """Сохранить ширины столбцов для файлового менеджера."""
    try:
        data = json.loads(request.body)
        widths = data.get('widths', {})
        category = data.get('category', 'EQUIPMENT')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Некорректные данные'}, status=400)

    user = request.user
    prefs = user.ui_preferences or {}
    fm_widths = prefs.get('fm_column_widths', {})
    fm_widths[category] = widths
    prefs['fm_column_widths'] = fm_widths
    user.ui_preferences = prefs
    user.save(update_fields=['ui_preferences'])
    return JsonResponse({'ok': True})
